#!/usr/bin/env python3
"""
evaluate_model.py — Model evaluation on validation / unseen datasets
IPI Antibody Developability Prediction Platform

Usage:
    python evaluate_model.py \
        --file   DS1_antiberta2-cssp_transformer_by_ipi.xlsx \
        --target psr_filter \
        --score  transformer_lm_antiberta2-cssp_ipi_psr_trainset_score

    python evaluate_model.py \
        --file   results.csv --target sec_filter \
        --score  xgboost_antiberta2-cssp_score --cost_fn 5.0

Outputs
-------
  1. Console — AUC + full threshold table
  2. Excel   — 3 sheets:
               • Summary (AUC, n, thresholds)
               • Threshold_Metrics (all methods × all metrics)
               • Predictions_with_OptimalLabels (original data +
                 optimal_label columns for each threshold method)
  3. TIFF    — Nature Biotechnology-style ROC figure (300 DPI)
  4. TIFF    — Score density plot by class (300 DPI)
  5. PNG     — Quick-view preview for both figures (200 DPI)
"""

import argparse, os
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from matplotlib.lines import Line2D

from sklearn.metrics import (
    roc_auc_score, roc_curve, auc,
    f1_score, fbeta_score,
    precision_score, recall_score,
    accuracy_score,
)


# ══════════════════════════════════════════════════════════════════════════════
# CORE METRICS
# ══════════════════════════════════════════════════════════════════════════════

def metrics_at(y_true, y_score, threshold):
    y_pred = (y_score >= threshold).astype(int)
    return {
        'threshold':   round(float(threshold), 4),
        'accuracy':    round(accuracy_score(y_true, y_pred),                        4),
        'f1':          round(f1_score(y_true, y_pred, zero_division=0),             4),
        'f2':          round(fbeta_score(y_true, y_pred, beta=2, zero_division=0),  4),
        'precision':   round(precision_score(y_true, y_pred, zero_division=0),      4),
        'sensitivity': round(recall_score(y_true, y_pred, zero_division=0),         4),
        'specificity': round(recall_score(y_true, y_pred, pos_label=0,
                                          zero_division=0),                          4),
        'recall_pass': round(recall_score(y_true, y_pred, zero_division=0),         4),
        'recall_fail': round(recall_score(y_true, y_pred, pos_label=0,
                                          zero_division=0),                          4),
        'pos_rate':    round(float(y_pred.mean()), 4),
    }


def find_thresholds(y_true, y_score, cost_fp=1.0, cost_fn=3.0):
    fpr, tpr, thr = roc_curve(y_true, y_score)
    spec     = 1.0 - fpr
    pos_rate = float(y_true.mean())

    # Youden J
    j_idx    = int(np.argmax(tpr + spec - 1))
    t_youden = float(thr[j_idx])

    # Cost-sensitive
    costs  = cost_fp * fpr * (1 - pos_rate) + cost_fn * (1 - tpr) * pos_rate
    t_cost = float(thr[int(np.argmin(costs))])

    # F-beta — use dense linspace sweep instead of roc_curve thresholds.
    # roc_curve thresholds are positioned at actual score values, which for
    # bimodal distributions (common on cross-lab transfer) can miss the true
    # optimal region. A 1000-point uniform grid is more reliable.
    smp = np.linspace(0.001, 0.999, 1000)

    def _sweep(beta):
        return [fbeta_score(y_true, (y_score>=t).astype(int),
                            beta=beta, zero_division=0) for t in smp]

    t_f1  = float(smp[int(np.argmax(_sweep(1.0)))])
    t_f2  = float(smp[int(np.argmax(_sweep(2.0)))])
    t_f05 = float(smp[int(np.argmax(_sweep(0.5)))])

    rec90_idx = np.where(tpr >= 0.90)[0]
    t_rec90   = float(thr[rec90_idx[0]]) if len(rec90_idx) else t_youden
    rec95_idx = np.where(tpr >= 0.95)[0]
    t_rec95   = float(thr[rec95_idx[0]]) if len(rec95_idx) else t_youden

    return {
        'Youden J (balanced)':                t_youden,
        f'Cost (FP={cost_fp}, FN={cost_fn})': t_cost,
        'F1 optimum':                         t_f1,
        'F2 (max class-1 recall)':            t_f2,
        'F0.5 (max class-1 precision)':       t_f05,
        'Sensitivity >= 90%':                 t_rec90,
        'Sensitivity >= 95%':                 t_rec95,
    }


# ══════════════════════════════════════════════════════════════════════════════
# SCORE DENSITY PLOT
# ══════════════════════════════════════════════════════════════════════════════
# SCORE DENSITY PLOT — Nature Biotech style
# ══════════════════════════════════════════════════════════════════════════════

def plot_score_density(y_true, y_score, thresh_dict,
                       out_tiff, label_col='', auc_val=None,
                       file_label=''):
    """
    Score distribution as bar histogram by class — Nature Biotech style.
    Pure histogram — no KDE overlay. See plot_kde_density() for smooth curves.
    """
    _nature_rcparams()

    s1 = y_score[y_true == 1];  n1 = len(s1)
    s0 = y_score[y_true == 0];  n0 = len(s0)
    COL1 = '#1f77b4'; COL0 = '#d62728'
    BINS = 25; EDGES = np.linspace(0, 1, BINS + 1)

    fig, ax = plt.subplots(figsize=(5.5, 3.2))
    legend_handles = []

    for s_cls, col, cls_label in [
        (s0, COL0, f'Class 0 — Fail  (n={n0:,})'),
        (s1, COL1, f'Class 1 — Pass  (n={n1:,})'),
    ]:
        if len(s_cls) < 1: continue
        counts, _ = np.histogram(s_cls, bins=EDGES)
        fracs = counts / counts.max()
        ax.bar(EDGES[:-1], fracs, width=(EDGES[1]-EDGES[0]),
               align='edge', color=col, alpha=0.55,
               edgecolor='white', linewidth=0.3, zorder=3)
        mn  = float(s_cls.mean());  med = float(np.median(s_cls))
        ax.axvline(mn,  color=col, lw=1.1, ls='--', alpha=0.9, zorder=5)
        ax.axvline(med, color=col, lw=1.1, ls=':',  alpha=0.9, zorder=5)
        ax.text(mn,  1.04,        f'\u03bc={mn:.2f}',   ha='center', va='bottom',
                fontsize=5.8, color=col)
        ax.text(med, 1.14, f'med={med:.2f}', ha='center', va='bottom',
                fontsize=5.8, color=col, style='italic')
        from matplotlib.lines import Line2D as _L
        legend_handles.append(_L([0],[0], color=col, lw=5, alpha=0.55, label=cls_label))

    t_youden = thresh_dict.get('Youden J (balanced)', 0.5)
    ax.axvline(t_youden, color='#333333', lw=1.2, ls='-.', zorder=6)
    from matplotlib.lines import Line2D as _L
    legend_handles += [
        _L([0],[0], color='#333333', lw=1.2, ls='-.', label=f'Youden  t={t_youden:.3f}'),
        _L([0],[0], color='#555555', lw=1.0, ls='--', label='Mean (\u03bc)'),
        _L([0],[0], color='#555555', lw=1.0, ls=':',  label='Median'),
    ]

    ax.set_xlim([0, 1]); ax.set_ylim([0, 1.28])
    ax.set_yticks([0, 0.5, 1.0])
    ax.yaxis.set_major_locator(ticker.FixedLocator([0, 0.5, 1.0]))
    ax.xaxis.set_major_locator(ticker.MultipleLocator(0.2))
    ax.tick_params(labelsize=6.5)
    ax.set_ylabel('Relative frequency\n(normalised per class to max=1)', fontsize=6.5, labelpad=4)
    _fl = file_label if file_label else label_col
    ax.set_xlabel(f'Predicted score\n{_fl}', fontsize=6, labelpad=6, color='#333333')
    _title = '  |  '.join(([label_col] if label_col else []) +
                           ([f'AUC = {auc_val:.3f}'] if auc_val is not None else []))
    ax.set_title(_title, fontsize=6.5, pad=5, color='#333333')
    ax.legend(handles=legend_handles, loc='upper left', bbox_to_anchor=(1.02, 1.0),
              fontsize=5.8, frameon=False, handlelength=1.5,
              labelspacing=0.4, borderpad=0)
    plt.tight_layout()
    _save_fig(fig, out_tiff, 'Histogram')


def plot_kde_density(y_true, y_score, thresh_dict,
                     out_tiff, label_col='', auc_val=None,
                     file_label=''):
    """
    Score distribution as KDE density curves by class — Nature Biotech style.
    Smooth curves only — no histogram bars. See plot_score_density() for bars.
    Requires scipy; falls back gracefully if absent.
    """
    _nature_rcparams()

    s1 = y_score[y_true == 1];  n1 = len(s1)
    s0 = y_score[y_true == 0];  n0 = len(s0)
    COL1 = '#1f77b4'; COL0 = '#d62728'
    x_grid = np.linspace(0, 1, 500)

    fig, ax = plt.subplots(figsize=(5.5, 3.2))
    legend_handles = []

    try:
        from scipy.stats import gaussian_kde
        _has_kde = True
    except ImportError:
        _has_kde = False

    for s_cls, col, cls_label in [
        (s0, COL0, f'Class 0 — Fail  (n={n0:,})'),
        (s1, COL1, f'Class 1 — Pass  (n={n1:,})'),
    ]:
        if len(s_cls) < 2: continue

        if _has_kde:
            try:
                kde  = gaussian_kde(s_cls, bw_method='scott')
                dens = kde(x_grid)
                dens = dens / dens.max()  # normalise to 1
                ax.plot(x_grid, dens, color=col, lw=1.5, zorder=4)
                ax.fill_between(x_grid, dens, alpha=0.12, color=col, zorder=3)
            except Exception:
                # fallback to step histogram if KDE fails
                counts, edges = np.histogram(s_cls, bins=25, range=(0,1))
                cx = 0.5*(edges[:-1]+edges[1:])
                ax.step(cx, counts/counts.max(), color=col, lw=1.5, where='mid')
        else:
            counts, edges = np.histogram(s_cls, bins=25, range=(0,1))
            cx = 0.5*(edges[:-1]+edges[1:])
            ax.step(cx, counts/counts.max(), color=col, lw=1.5, where='mid')

        mn  = float(s_cls.mean());  med = float(np.median(s_cls))
        ax.axvline(mn,  color=col, lw=1.1, ls='--', alpha=0.9, zorder=5)
        ax.axvline(med, color=col, lw=1.1, ls=':',  alpha=0.9, zorder=5)
        ax.text(mn,  1.04, f'\u03bc={mn:.2f}',   ha='center', va='bottom',
                fontsize=5.8, color=col)
        ax.text(med, 1.14, f'med={med:.2f}', ha='center', va='bottom',
                fontsize=5.8, color=col, style='italic')
        from matplotlib.lines import Line2D as _L
        legend_handles.append(_L([0],[0], color=col, lw=2.0, label=cls_label))

    t_youden = thresh_dict.get('Youden J (balanced)', 0.5)
    ax.axvline(t_youden, color='#333333', lw=1.2, ls='-.', zorder=6)
    from matplotlib.lines import Line2D as _L
    legend_handles += [
        _L([0],[0], color='#333333', lw=1.2, ls='-.', label=f'Youden  t={t_youden:.3f}'),
        _L([0],[0], color='#555555', lw=1.0, ls='--', label='Mean (\u03bc)'),
        _L([0],[0], color='#555555', lw=1.0, ls=':',  label='Median'),
    ]

    ax.set_xlim([0, 1]); ax.set_ylim([0, 1.28])
    ax.set_yticks([0, 0.5, 1.0])
    ax.yaxis.set_major_locator(ticker.FixedLocator([0, 0.5, 1.0]))
    ax.xaxis.set_major_locator(ticker.MultipleLocator(0.2))
    ax.tick_params(labelsize=6.5)
    ax.set_ylabel('Density  (normalised to max=1)', fontsize=6.5, labelpad=4)
    _fl = file_label if file_label else label_col
    ax.set_xlabel(f'Predicted score\n{_fl}', fontsize=6, labelpad=6, color='#333333')
    _title = '  |  '.join(([label_col] if label_col else []) +
                           ([f'AUC = {auc_val:.3f}'] if auc_val is not None else []))
    ax.set_title(_title, fontsize=6.5, pad=5, color='#333333')
    ax.legend(handles=legend_handles, loc='upper left', bbox_to_anchor=(1.02, 1.0),
              fontsize=5.8, frameon=False, handlelength=1.5,
              labelspacing=0.4, borderpad=0)
    plt.tight_layout()
    _save_fig(fig, out_tiff, 'KDE density')


def _nature_rcparams():
    """Apply Nature Biotech rcParams — call at start of every plot function."""
    plt.rcParams.update({
        'font.family':      'sans-serif',
        'font.sans-serif':  ['Arial', 'Helvetica', 'DejaVu Sans'],
        'font.size':         7,
        'axes.linewidth':    0.8,
        'axes.spines.top':   False,
        'axes.spines.right': False,
        'xtick.major.width': 0.8,
        'ytick.major.width': 0.8,
        'xtick.major.size':  3,
        'ytick.major.size':  3,
        'xtick.direction':   'out',
        'ytick.direction':   'out',
        'pdf.fonttype':      42,
        'ps.fonttype':       42,
    })


def _save_fig(fig, out_tiff, label=''):
    """Save figure as TIFF (300 DPI) + PNG preview (200 DPI), then close."""
    out_png = out_tiff.replace('.tiff', '_preview.png').replace('.tif', '_preview.png')
    fig.savefig(out_tiff, dpi=300, format='tiff', bbox_inches='tight', facecolor='white')
    fig.savefig(out_png,  dpi=200, bbox_inches='tight', facecolor='white')
    plt.close()
    plt.rcdefaults()
    print(f"[plot] {label} TIFF 300 DPI → {out_tiff}")
    print(f"[plot] {label} PNG  200 DPI → {out_png}")










def plot_roc_nature(y_true, y_score, auc_val, thresh_dict,
                    out_tiff, label='', cost_fn=3.0, cost_fp=1.0):
    """
    ROC curve meeting Nature Biotechnology / Nature Methods figure standards:
      - 3.5 x 3.5 inches  (single column width)
      - 300 DPI TIFF + 200 DPI PNG preview
      - Arial / Helvetica, 7 pt body, clean white background
      - No top/right spines, outward ticks, no grid
      - AUC value annotated inside the plot
      - Operating points labelled with threshold value only
      - Legend outside to the right (no frame)
    """
    plt.rcParams.update({
        'font.family':      'sans-serif',
        'font.sans-serif':  ['Arial', 'Helvetica', 'DejaVu Sans'],
        'font.size':         7,
        'axes.linewidth':    0.8,
        'axes.spines.top':   False,
        'axes.spines.right': False,
        'xtick.major.width': 0.8,
        'ytick.major.width': 0.8,
        'xtick.major.size':  3,
        'ytick.major.size':  3,
        'xtick.direction':   'out',
        'ytick.direction':   'out',
        'pdf.fonttype':      42,
        'ps.fonttype':       42,
    })

    _nature_rcparams()
    fpr_curve, tpr_curve, _ = roc_curve(y_true, y_score)

    # Wide figure so legend fits outside without clipping the square plot area
    fig, ax = plt.subplots(figsize=(5.0, 3.5))
    ax.set_aspect('equal')   # ← perfect square plot area (Nature standard)

    # ROC curve — dark navy, 1.5 pt
    ax.plot(fpr_curve, tpr_curve,
            color='#1a1a2e', lw=1.5, zorder=3, solid_capstyle='round')

    # Diagonal reference — light grey dashes
    ax.plot([0, 1], [0, 1],
            color='#aaaaaa', lw=0.8, linestyle='--',
            dashes=(4, 3), zorder=2)

    # Shaded area under curve — very subtle
    ax.fill_between(fpr_curve, tpr_curve,
                    alpha=0.06, color='#1a1a2e', zorder=1)

    # AUC annotation — bottom right, italic
    ax.text(0.97, 0.05,
            f'AUC = {auc_val:.3f}',
            transform=ax.transAxes,
            ha='right', va='bottom',
            fontsize=7.5, fontstyle='italic',
            color='#1a1a2e', fontweight='semibold')

    # Operating points
    # Clean display names without redundant parenthetical repetition
    _styles = {
        'Youden J (balanced)':                ('Youden J',                  '#d62728', 'o', 6.5),
        f'Cost (FP={cost_fp}, FN={cost_fn})': (f'Cost  FN={cost_fn}\u00d7FP','#ff7f0e', 's', 5.5),
        'F1 optimum':                         ('F1 optimum',                '#1f77b4', '^', 6.0),
        'F2 (max class-1 recall)':            ('F2  (max recall)',           '#9467bd', 'D', 5.0),
        'F0.5 (max class-1 precision)':       ('F0.5  (max precision)',      '#8c564b', 'v', 5.5),
        'Sensitivity >= 90%':                 ('Sensitivity \u226590%',      '#17becf', 'P', 6.0),
        'Sensitivity >= 95%':                 ('Sensitivity \u226595%',      '#bcbd22', 'X', 6.0),
        't = 0.500':                          ('t = 0.500',                  '#7f7f7f', '*', 9.0),
    }

    legend_handles = []
    all_methods = list(thresh_dict.items()) + [('t = 0.500', 0.5)]

    for method, t in all_methods:
        if method not in _styles:
            continue   # method not in display list — skip
        disp, col, mk, sz = _styles[method]
        m    = metrics_at(y_true, y_score, t)
        fpr_ = 1.0 - m['specificity']
        sens_= m['sensitivity']
        ax.scatter(fpr_, sens_,
                   color=col, marker=mk, s=sz**2,
                   zorder=5, edgecolors='white', linewidths=0.5)
        legend_handles.append(
            Line2D([0], [0], marker=mk, color='w',
                   markerfacecolor=col, markersize=sz * 0.85,
                   markeredgecolor='white', markeredgewidth=0.3,
                   label=f'{disp}  (t={t:.3f})'))

    # Axes formatting
    ax.set_xlim([0, 1])
    ax.set_ylim([0, 1])
    ax.set_xlabel('1 \u2212 Specificity  (False Positive Rate)',
                  fontsize=7, labelpad=4)
    ax.set_ylabel('Sensitivity  (True Positive Rate)',
                  fontsize=7, labelpad=4)
    ax.xaxis.set_major_locator(ticker.MultipleLocator(0.2))
    ax.yaxis.set_major_locator(ticker.MultipleLocator(0.2))
    ax.xaxis.set_minor_locator(ticker.MultipleLocator(0.1))
    ax.yaxis.set_minor_locator(ticker.MultipleLocator(0.1))
    ax.tick_params(which='major', labelsize=6.5, length=3)
    ax.tick_params(which='minor', length=1.5, color='#aaaaaa')

    # Title — concise
    short_label = (label[:52] + '\u2026') if len(label) > 52 else label
    ax.set_title(short_label, fontsize=6.5, pad=5, color='#333333')

    # Legend — outside right, no frame
    ax.legend(
        handles       = legend_handles,
        loc           = 'upper left',
        bbox_to_anchor= (1.03, 1.01),
        fontsize      = 5.8,
        frameon       = False,
        handlelength  = 1.2,
        handletextpad = 0.5,
        borderpad     = 0,
        labelspacing  = 0.35,
    )

    plt.tight_layout()
    _save_fig(fig, out_tiff, 'ROC')

# ══════════════════════════════════════════════════════════════════════════════
# MAIN EVALUATE
# ══════════════════════════════════════════════════════════════════════════════

def evaluate(file, target, score_col, cost_fp=1.0, cost_fn=3.0,
             out=None, test_target=None,
             model_type=None, lm=None, db_stem=None,
             dataset_name=None):
    """
    Main evaluation function.  Returns (auc_val, df_metrics, thresh_dict).

    test_target : str or None
        Ground-truth column name in the test file when it differs from --target.
        Example: --target psr_filter (model was trained on this)
                 --test_target psr_anno (unseen dataset uses this column name)
        If None (default): uses --target if that column exists in the file.
    """

    # ── Load ──────────────────────────────────────────────────────────────────
    ext = os.path.splitext(file)[1].lower()
    if ext in ('.xlsx', '.xls'):
        df = pd.read_excel(file)
    elif ext == '.csv':
        df = pd.read_csv(file)
    else:
        raise ValueError(f"Unsupported format: {ext}")

    for col in [score_col]:
        if col not in df.columns:
            raise ValueError(f"Score column '{score_col}' not found.\nAvailable: {list(df.columns)}")

    # ── Resolve ground-truth column ───────────────────────────────────────────
    # Priority: test_target > target (if present in file)
    # If neither found → skip evaluation (called from auto_predict context)
    _label_col = None
    if test_target is not None:
        if test_target not in df.columns:
            raise ValueError(
                f"--test_target '{test_target}' not found in file.\n"
                f"Available: {list(df.columns)}"
            )
        _label_col = test_target
        if test_target != target:
            print(f"[eval] Using test_target='{test_target}' as ground-truth "
                  f"(model was trained on '{target}')")
    elif target in df.columns:
        _label_col = target
    else:
        raise ValueError(
            f"No ground-truth column found.\n"
            f"  --target '{target}' not in file.\n"
            f"  Pass --test_target <colname> if your test file uses a different label name.\n"
            f"  Available columns: {list(df.columns)}"
        )

    df_eval = df[[_label_col, score_col]].dropna()
    y       = df_eval[_label_col].astype(int).values
    s       = df_eval[score_col].astype(float).values
    s_full  = df[score_col].fillna(0.0).astype(float).values

    if len(np.unique(y)) < 2:
        raise ValueError("Only one class present — cannot compute AUC.")

    n        = len(y)
    pos_rate = float(y.mean())
    auc_val  = roc_auc_score(y, s)

    # ── Output stems ──────────────────────────────────────────────────────────
    stem     = os.path.splitext(out or file)[0].replace(' ', '_')
    out_xlsx = f"{stem}_eval_{_label_col}.xlsx"
    out_tiff = f"{stem}_roc_{_label_col}.tiff"

    # ── Thresholds ────────────────────────────────────────────────────────────
    thresh_dict = find_thresholds(y, s, cost_fp=cost_fp, cost_fn=cost_fn)

    rows = {'t = 0.500': {'method': 't = 0.500', **metrics_at(y, s, 0.500)}}
    for method, t in thresh_dict.items():
        rows[method] = {'method': method, **metrics_at(y, s, t)}

    # ── Console ───────────────────────────────────────────────────────────────
    W = 76
    print(f"\n{'═'*W}")
    print(f"  EVALUATION REPORT")
    print(f"  File        : {os.path.basename(file)}")
    print(f"  Model target: {target}  |  Score : {score_col}")
    print(f"  Label used  : {_label_col}"
          + (f"  (remapped from --test_target)" if _label_col != target else ""))
    print(f"  n={n:,}  pos_rate={pos_rate:.1%}  "
          f"(class 1={int(y.sum()):,}  class 0={int((1-y).sum()):,})")
    print(f"  Score  : min={s.min():.4f}  max={s.max():.4f}  mean={s.mean():.4f}")
    print(f"{'─'*W}")
    print(f"  AUC-ROC = {auc_val:.4f}")
    print(f"{'─'*W}")

    col_w = [28, 7, 8, 7, 7, 7, 7, 7, 7, 7, 7]
    header = (f"  {'Method':<{col_w[0]}}  {'Thresh':>{col_w[1]}}  "
              f"{'Acc':>{col_w[2]}}  {'F1':>{col_w[3]}}  {'F2':>{col_w[4]}}  "
              f"{'Prec':>{col_w[5]}}  {'Sens':>{col_w[6]}}  {'Spec':>{col_w[7]}}  "
              f"{'Rec1':>{col_w[8]}}  {'Rec0':>{col_w[9]}}  {'Pos%':>{col_w[10]}}")
    print(header)
    print("  " + "─" * (len(header) - 2))
    for method, r in rows.items():
        flag = "  <--" if method == 'Youden J (balanced)' else ""
        print(f"  {method:<{col_w[0]}}  {r['threshold']:>{col_w[1]}.4f}  "
              f"{r['accuracy']:>{col_w[2]}.4f}  {r['f1']:>{col_w[3]}.4f}  "
              f"{r['f2']:>{col_w[4]}.4f}  {r['precision']:>{col_w[5]}.4f}  "
              f"{r['sensitivity']:>{col_w[6]}.4f}  {r['specificity']:>{col_w[7]}.4f}  "
              f"{r['recall_pass']:>{col_w[8]}.4f}  {r['recall_fail']:>{col_w[9]}.4f}  "
              f"{r['pos_rate']:>{col_w[10]}.1%}{flag}")

    youden_t = thresh_dict['Youden J (balanced)']
    cost_t   = thresh_dict[f'Cost (FP={cost_fp}, FN={cost_fn})']
    t_f2     = thresh_dict['F2 (max class-1 recall)']
    t_f05    = thresh_dict['F0.5 (max class-1 precision)']
    t_s90    = thresh_dict['Sensitivity >= 90%']
    print(f"{'─'*W}")
    print(f"  Class 1 = non-polyreactive (PASS)  |  Class 0 = polyreactive (FAIL)")
    print(f"{'─'*W}")
    print(f"  Recommended thresholds:")
    print(f"    Balanced FP/FN               → Youden     t={youden_t:.4f}  equal confidence Pass/Fail")
    print(f"    Max class-1 recall           → F2         t={t_f2:.4f}  miss fewest good antibodies")
    print(f"    Max class-1 precision        → F0.5       t={t_f05:.4f}  fewest polyreactives advanced")
    print(f"    Guarantee 90% good recovered → Sens\u226590%  t={t_s90:.4f}  standard screening cutoff")
    print(f"    Cost-aware (miss good={cost_fn}\u00d7bad) → Cost      t={cost_t:.4f}  quantified cost tradeoff")
    print(f"{'═'*W}\n")

    # ── Add optimal_label columns to dataframe ─────────────────────────────
    # Column naming convention:
    #   score_col = '{model}_{lm}_{db}_score'
    #   → model_prefix = '{model}_{lm}_{db}'
    #   → new col = '{model_prefix}_optimallabel'  (Youden)
    #               '{model_prefix}_costlabel'     (cost-sensitive)
    #               '{model_prefix}_f1label'       (F1 optimum)
    #               '{model_prefix}_f2label'       (f2label)
    #               '{model_prefix}_sens90label'   (sensitivity >= 90%)
    model_prefix = score_col.replace('_score', '')

    label_cols = {
        'optimallabel': ('Youden J (balanced)',                youden_t),
        'costlabel':    (f'Cost (FP={cost_fp}, FN={cost_fn})', cost_t),
        'f1label':      ('F1 optimum',                         thresh_dict['F1 optimum']),
        'f2label':      ('F2 (max class-1 recall)',            thresh_dict['F2 (max class-1 recall)']),
        'f05label':     ('F0.5 (max class-1 precision)',       thresh_dict['F0.5 (max class-1 precision)']),
        'sens90label':  ('Sensitivity >= 90%',                 thresh_dict['Sensitivity >= 90%']),
    }

    for suffix, (method_name, t) in label_cols.items():
        col_name = f"{model_prefix}_{suffix}"
        df[col_name] = (s_full >= t).astype(int)

    # ── Excel ─────────────────────────────────────────────────────────────────
    metrics_cols = ['method','threshold','accuracy','f1','f2','precision',
                    'sensitivity','specificity','recall_pass','recall_fail','pos_rate']
    df_metrics   = pd.DataFrame(list(rows.values()))[metrics_cols]
    df_metrics.columns = [
        'Method','Threshold','Accuracy','F1','F2','Precision',
        'Sensitivity','Specificity','Recall (class 1)',
        'Recall (class 0 / Fail)','Pos Rate'
    ]

    summary_data = {
        'Metric': ['AUC-ROC','n','pos_rate','class_1_n','class_0_n',
                   'score_min','score_max','score_mean',
                   'threshold_youden','threshold_cost','threshold_f1','threshold_f2'],
        'Value':  [round(auc_val,4), n, round(pos_rate,4),
                   int(y.sum()), int((1-y).sum()),
                   round(float(s.min()),4), round(float(s.max()),4),
                   round(float(s.mean()),4),
                   round(youden_t,4), round(cost_t,4),
                   round(thresh_dict['F1 optimum'],4),
                   round(thresh_dict['F2 (max class-1 recall)'],4)],
    }

    with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
        pd.DataFrame(summary_data).to_excel(
            writer, sheet_name='Summary', index=False)
        df_metrics.to_excel(
            writer, sheet_name='Threshold_Metrics', index=False)
        df.to_excel(
            writer, sheet_name='Predictions_with_OptimalLabels', index=False)

        wb = writer.book
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        _thin = Side(style='thin', color='D0D0D0')
        _border = Border(bottom=_thin)

        for sn in ['Summary', 'Threshold_Metrics', 'Predictions_with_OptimalLabels']:
            ws = wb[sn]
            # Headers — bold, no colour fill, bottom border only
            for cell in ws[1]:
                cell.font      = Font(bold=True, name='Arial', size=9)
                cell.fill      = PatternFill(fill_type=None)  # no fill
                cell.alignment = Alignment(horizontal='center')
                cell.border    = _border
            # Youden row — light grey text weight only, no fill
            if sn == 'Threshold_Metrics':
                for row in ws.iter_rows(min_row=2):
                    if row[0].value and 'Youden' in str(row[0].value):
                        for cell in row:
                            cell.font = Font(bold=True, name='Arial', size=9)
            # Data rows — regular font
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.font is None or not cell.font.bold:
                        cell.font = Font(name='Arial', size=9)
            # Auto width
            for col in ws.columns:
                max_len = max(len(str(c.value or '')) for c in col) + 2
                ws.column_dimensions[
                    get_column_letter(col[0].column)].width = min(max_len, 40)

    print(f"[eval] Excel  → {out_xlsx}")
    print(f"[eval]   Sheet 1: Summary")
    print(f"[eval]   Sheet 2: Threshold_Metrics")
    print(f"[eval]   Sheet 3: Predictions_with_OptimalLabels")
    print(f"[eval]   New optimal_label columns added:")
    for suffix, (method_name, t) in label_cols.items():
        print(f"[eval]     {model_prefix}_{suffix}  (t={t:.4f}, method={method_name})")

    # ── Nature Biotech ROC ────────────────────────────────────────────────────
    _label = (f"{os.path.basename(file).replace('.xlsx','').replace('.csv','')}"
              f" | {_label_col}")
    plot_roc_nature(y, s, auc_val, thresh_dict,
                    out_tiff=out_tiff, label=_label,
                    cost_fn=cost_fn, cost_fp=cost_fp)

    # ── Score histogram (figure 2) ────────────────────────────────────────────
    # ── Reconstruct clean dataset name ────────────────────────────────────────
    # When called from auto_predict, 'file' is the prediction OUTPUT file:
    #   e.g. ipiab202603_pred_psr_filter_antiberta2-cssp_transformer_lm_ipi_psr_trainset.xlsx
    # We need just the original input stem: 'ipiab202603'
    #
    # Strategy (in order of reliability):
    #   1. dataset_name passed explicitly → use it directly
    #   2. model components known → strip the prediction suffix from filename
    #   3. fallback → use raw stem (may include model info but beats garbled label)
    _raw_stem = os.path.splitext(os.path.basename(file))[0]

    if dataset_name:
        # Explicit — most reliable (passed from auto_predict)
        _dataset_name = dataset_name
    elif model_type and lm and db_stem:
        # Strip the known suffix: _pred_{target}_{lm}_{model_type}_{db_stem}
        # Also try _eval_{target} suffix (when called on eval output file)
        _suffix = f"_pred_{_label_col}_{lm}_{model_type}_{db_stem}"
        _suffix_alt = f"_pred_{target}_{lm}_{model_type}_{db_stem}"
        if _raw_stem.endswith(_suffix):
            _dataset_name = _raw_stem[:-len(_suffix)]
        elif _raw_stem.endswith(_suffix_alt):
            _dataset_name = _raw_stem[:-len(_suffix_alt)]
        else:
            # Try stripping just _pred_{target} prefix
            _short_suffix = f"_pred_{_label_col}"
            _dataset_name = (_raw_stem[:-len(_short_suffix)]
                             if _raw_stem.endswith(_short_suffix) else _raw_stem)
    else:
        _dataset_name = _raw_stem

    # Build file_label: {dataset}_pred_by_{db_stem}_{model_type}_{lm}
    if model_type and lm and db_stem:
        _file_label = f"{_dataset_name}_pred_by_{db_stem}_{model_type}_{lm}"
    else:
        # Standalone CLI path — parse from score_col
        _known_models = ['transformer_lm', 'transformer_onehot',
                         'cnn', 'xgboost', 'mlp', 'rf', 'svm']
        _known_lms    = ['antiberta2-cssp', 'antiberta2', 'ablang2',
                         'esm2', 'onehot_vh', 'onehot', 'protbert']
        _remainder = score_col.replace('_score', '')
        _mt, _lm_found, _db = '', '', _remainder
        for _m in _known_models:
            if _remainder.startswith(_m + '_'):
                _mt = _m; _remainder = _remainder[len(_m)+1:]; break
        for _l in _known_lms:
            if _remainder.startswith(_l + '_') or _remainder == _l:
                _lm_found = _l
                _db = _remainder[len(_l)+1:] if '_' in _remainder[len(_l):] else ''
                break
        _file_label = f"{_dataset_name}_pred_by_{_db}_{_mt}_{_lm_found}"

    print(f"[eval] File label : {_file_label}")
    out_hist_tiff = out_tiff.replace('_roc_', '_histogram_')
    plot_score_density(y, s, thresh_dict,
                       out_tiff   = out_hist_tiff,
                       label_col  = _label_col,
                       auc_val    = auc_val,
                       file_label = _file_label)

    # ── KDE density (figure 3) ────────────────────────────────────────────────
    out_kde_tiff = out_tiff.replace('_roc_', '_kde_')
    plot_kde_density(y, s, thresh_dict,
                     out_tiff   = out_kde_tiff,
                     label_col  = _label_col,
                     auc_val    = auc_val,
                     file_label = _file_label)

    return auc_val, df_metrics, thresh_dict


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Evaluate model: AUC + optimal thresholds + Nature-style ROC',
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('--file',        required=True)
    parser.add_argument('--target',      required=True,
                        help='Label column name used during training (e.g. psr_filter)')
    parser.add_argument('--score',       required=True)
    parser.add_argument('--test_target', default=None,
                        help='Label column name in the test file if different from --target '
                             '(e.g. psr_anno, psr_filter2). '
                             'Default: uses --target if found in the file.')
    parser.add_argument('--cost_fp', type=float, default=1.0)
    parser.add_argument('--cost_fn', type=float, default=3.0)
    parser.add_argument('--out',     default=None)

    args = parser.parse_args()
    evaluate(
        file        = args.file,
        target      = args.target,
        score_col   = args.score,
        cost_fp     = args.cost_fp,
        cost_fn     = args.cost_fn,
        out         = args.out,
        test_target = args.test_target,
    )